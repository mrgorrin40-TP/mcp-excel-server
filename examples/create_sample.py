"""Create sample Excel file for testing."""

from openpyxl import Workbook
from pathlib import Path


def create_sample_file(output_path: str = "examples/data/sample.xlsx"):
    """Create a sample Excel file with test data."""
    wb = Workbook()
    
    # Sales sheet
    ws_sales = wb.active
    ws_sales.title = "Sales"
    
    # Headers
    ws_sales["A1"] = "Product"
    ws_sales["B1"] = "Region"
    ws_sales["C1"] = "Quantity"
    ws_sales["D1"] = "Price"
    ws_sales["E1"] = "Total"
    
    # Data
    data = [
        ["Widget A", "North", 100, 29.99],
        ["Widget B", "South", 50, 49.99],
        ["Widget C", "East", 75, 39.99],
        ["Widget A", "West", 120, 29.99],
        ["Widget B", "North", 80, 49.99],
        ["Widget C", "South", 60, 39.99],
        ["Widget A", "East", 90, 29.99],
        ["Widget B", "West", 110, 49.99],
    ]
    
    for i, row in enumerate(data, start=2):
        ws_sales[f"A{i}"] = row[0]
        ws_sales[f"B{i}"] = row[1]
        ws_sales[f"C{i}"] = row[2]
        ws_sales[f"D{i}"] = row[3]
        ws_sales[f"E{i}"] = f"=C{i}*D{i}"
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary["A1"] = "Region"
    ws_summary["B1"] = "Total Sales"
    
    ws_summary["A2"] = "North"
    ws_summary["B2"] = "=SUMIF(Sales!B:B,A2,Sales!E:E)"
    
    ws_summary["A3"] = "South"
    ws_summary["B3"] = "=SUMIF(Sales!B:B,A3,Sales!E:E)"
    
    ws_summary["A4"] = "East"
    ws_summary["B4"] = "=SUMIF(Sales!B:B,A4,Sales!E:E)"
    
    ws_summary["A5"] = "West"
    ws_summary["B5"] = "=SUMIF(Sales!B:B,A5,Sales!E:E)"
    
    ws_summary["A7"] = "Grand Total"
    ws_summary["B7"] = "=SUM(B2:B5)"
    
    # Customers sheet
    ws_customers = wb.create_sheet("Customers")
    
    ws_customers["A1"] = "Name"
    ws_customers["B1"] = "Email"
    ws_customers["C1"] = "Region"
    ws_customers["D1"] = "Status"
    
    customers = [
        ["Alice Johnson", "alice@example.com", "North", "Active"],
        ["Bob Smith", "bob@example.com", "South", "Active"],
        ["Charlie Brown", "charlie@example.com", "East", "Inactive"],
        ["Diana Ross", "diana@example.com", "West", "Active"],
        ["Eve Davis", "eve@example.com", "North", "Active"],
    ]
    
    for i, row in enumerate(customers, start=2):
        for j, value in enumerate(row):
            ws_customers.cell(row=i, column=j + 1, value=value)
    
    # Save file
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"Sample file created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_sample_file()
