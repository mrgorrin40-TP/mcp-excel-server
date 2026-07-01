"""MCP resources for Excel file metadata."""

from typing import Any


def get_workbook_metadata(file_path: str) -> dict[str, Any]:
    """Get workbook metadata as MCP resource.

    Args:
        file_path: Absolute path to Excel file

    Returns:
        Dictionary with workbook metadata
    """
    from pathlib import Path

    from ..backends.factory import create_backend

    try:
        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        backend = create_backend(file_path)
        backend.open(file_path)

        sheet_names = backend.get_sheet_names()
        sheets_info = []

        for sheet_name in sheet_names:
            try:
                ws = backend.get_sheet(sheet_name)
                sheets_info.append({
                    "name": sheet_name,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                })
            except Exception:
                sheets_info.append({"name": sheet_name, "error": "Could not read"})

        backend.close()

        return {
            "uri": f"excel:///{path.name}",
            "name": path.name,
            "path": str(path.resolve()),
            "size_kb": round(path.stat().st_size / 1024, 2),
            "sheets": sheets_info,
            "total_sheets": len(sheets_info),
        }
    except Exception as e:
        return {"error": str(e)}


def get_sheet_metadata(file_path: str, sheet_name: str) -> dict[str, Any]:
    """Get sheet metadata as MCP resource.

    Args:
        file_path: Absolute path to Excel file
        sheet_name: Worksheet name

    Returns:
        Dictionary with sheet metadata
    """
    from pathlib import Path

    from ..backends.factory import create_backend

    try:
        path = Path(file_path)
        backend = create_backend(file_path)
        backend.open(file_path)

        ws = backend.get_sheet(sheet_name)

        # Get headers
        headers = []
        for col in range(1, (ws.max_column or 0) + 1):
            from openpyxl.utils import get_column_letter
            cell = ws[f"{get_column_letter(col)}1"]
            headers.append({
                "name": str(cell.value) if cell.value else f"Column {col}",
                "letter": get_column_letter(col),
                "type": backend.get_cell_type(sheet_name, f"{get_column_letter(col)}1"),
            })

        backend.close()

        return {
            "uri": f"excel:///{path.name}/{sheet_name}",
            "name": sheet_name,
            "rows": ws.max_row or 0,
            "columns": ws.max_column or 0,
            "headers": headers,
            "used_range": backend.get_used_range(sheet_name),
        }
    except Exception as e:
        return {"error": str(e)}
