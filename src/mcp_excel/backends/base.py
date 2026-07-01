"""Abstract base class for Excel backends."""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any


class ExcelBackend(ABC):
    """Abstract base class for Excel file operations."""

    @abstractmethod
    def open(self, file_path: str | Path) -> None:
        """Open an Excel file."""
        pass

    @abstractmethod
    def close(self) -> None:
        """Close the current workbook."""
        pass

    @abstractmethod
    def save(self, file_path: str | Path | None = None) -> None:
        """Save the workbook."""
        pass

    @abstractmethod
    def get_sheet_names(self) -> list[str]:
        """Get list of all sheet names."""
        pass

    @abstractmethod
    def get_sheet(self, sheet_name: str) -> Any:
        """Get a worksheet by name."""
        pass

    @abstractmethod
    def create_sheet(self, sheet_name: str) -> Any:
        """Create a new worksheet."""
        pass

    @abstractmethod
    def delete_sheet(self, sheet_name: str) -> None:
        """Delete a worksheet."""
        pass

    @abstractmethod
    def read_cell(self, sheet_name: str, cell: str) -> Any:
        """Read a single cell value."""
        pass

    @abstractmethod
    def write_cell(self, sheet_name: str, cell: str, value: Any) -> None:
        """Write a value to a cell."""
        pass

    @abstractmethod
    def read_range(self, sheet_name: str, range_str: str) -> list[list[Any]]:
        """Read a range of cells."""
        pass

    @abstractmethod
    def write_range(self, sheet_name: str, range_str: str, values: list[list[Any]]) -> None:
        """Write values to a range."""
        pass

    @abstractmethod
    def get_used_range(self, sheet_name: str) -> str:
        """Get the used range of a worksheet."""
        pass

    @abstractmethod
    def get_cell_type(self, sheet_name: str, cell: str) -> str:
        """Get the data type of a cell."""
        pass

    # Sheet metadata methods (for tools that need row/column counts)

    @abstractmethod
    def get_max_row(self, sheet_name: str) -> int:
        """Get the maximum row number with data in a worksheet."""
        pass

    @abstractmethod
    def get_max_column(self, sheet_name: str) -> int:
        """Get the maximum column number with data in a worksheet."""
        pass

    def get_cell_value(self, sheet_name: str, cell: str) -> Any:
        """Get the value of a cell (convenience method).

        Default implementation uses read_cell.
        Override for better performance if needed.
        """
        return self.read_cell(sheet_name, cell)

    def get_row_values(self, sheet_name: str, row: int) -> list[Any]:
        """Get all values in a row.

        Default implementation uses read_range.
        Override for better performance if needed.
        """
        from openpyxl.utils import get_column_letter

        max_col = self.get_max_column(sheet_name)
        if max_col == 0:
            return []
        range_str = f"A{row}:{get_column_letter(max_col)}{row}"
        result = self.read_range(sheet_name, range_str)
        return result[0] if result else []

    def get_column_values(self, sheet_name: str, column: str) -> list[Any]:
        """Get all values in a column.

        Default implementation uses read_range.
        Override for better performance if needed.
        """
        max_row = self.get_max_row(sheet_name)
        if max_row == 0:
            return []
        range_str = f"{column}1:{column}{max_row}"
        result = self.read_range(sheet_name, range_str)
        return [row[0] if row else None for row in result]

    # VBA-specific methods (optional - not all backends support VBA)

    def has_macros(self) -> bool:
        """Check if workbook contains VBA macros.

        Default implementation returns False.
        Override in backends that support VBA.
        """
        return False

    def list_vba_modules(self) -> list[dict[str, Any]]:
        """List all VBA modules in the workbook.

        Default implementation returns empty list.
        Override in backends that support VBA.
        """
        return []

    def get_vba_code(self, module_name: str) -> str:
        """Get VBA source code from a module.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def set_vba_code(self, module_name: str, code: str) -> None:
        """Set/replace VBA source code in a module.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def add_vba_module(self, name: str, code: str = "", module_type: str = "standard") -> None:
        """Add a new VBA module.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def delete_vba_module(self, name: str) -> None:
        """Delete a VBA module.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def rename_vba_module(self, old_name: str, new_name: str) -> None:
        """Rename a VBA module.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def run_macro(self, macro_name: str, *args: Any) -> Any:
        """Execute a VBA macro by name.

        Default implementation raises ValueError.
        Override in backends that support VBA.
        """
        raise ValueError("VBA not supported by this backend")

    def list_macros(self) -> list[dict[str, Any]]:
        """List all Sub/Function procedures in VBA project.

        Default implementation returns empty list.
        Override in backends that support VBA.
        """
        return []
