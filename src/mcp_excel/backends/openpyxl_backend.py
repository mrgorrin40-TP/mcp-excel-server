"""Openpyxl backend for Excel file operations."""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .base import ExcelBackend

logger = logging.getLogger(__name__)


class OpenpyxlBackend(ExcelBackend):
    """Excel backend using openpyxl library."""

    def __init__(self) -> None:
        self._workbook: Workbook | None = None
        self._file_path: Path | None = None

    def open(self, file_path: str | Path) -> None:
        """Open an Excel file."""
        self._file_path = Path(file_path)

        if not self._file_path.exists():
            raise FileNotFoundError(f"File not found: {self._file_path}")

        if not self._file_path.suffix.lower() == ".xlsx":
            raise ValueError(f"Unsupported file format: {self._file_path.suffix}")

        logger.info("Opening workbook: %s", self._file_path)
        self._workbook = load_workbook(
            self._file_path,
            read_only=False,
            data_only=False,
        )

    def close(self) -> None:
        """Close the current workbook."""
        if self._workbook:
            logger.info("Closing workbook")
            self._workbook.close()
            self._workbook = None
            self._file_path = None

    def save(self, file_path: str | Path | None = None) -> None:
        """Save the workbook."""
        if not self._workbook:
            raise ValueError("No workbook open")

        save_path = Path(file_path) if file_path else self._file_path
        if not save_path:
            raise ValueError("No file path specified")

        logger.info("Saving workbook to: %s", save_path)
        self._workbook.save(save_path)
        self._file_path = save_path

    def get_sheet_names(self) -> list[str]:
        """Get list of all sheet names."""
        if not self._workbook:
            raise ValueError("No workbook open")

        return list(self._workbook.sheetnames)

    def get_sheet(self, sheet_name: str) -> Any:
        """Get a worksheet by name."""
        if not self._workbook:
            raise ValueError("No workbook open")

        if sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        return self._workbook[sheet_name]

    def create_sheet(self, sheet_name: str) -> Any:
        """Create a new worksheet."""
        if not self._workbook:
            raise ValueError("No workbook open")

        if sheet_name in self._workbook.sheetnames:
            raise ValueError(f"Sheet already exists: {sheet_name}")

        logger.info("Creating sheet: %s", sheet_name)
        return self._workbook.create_sheet(title=sheet_name)

    def delete_sheet(self, sheet_name: str) -> None:
        """Delete a worksheet."""
        if not self._workbook:
            raise ValueError("No workbook open")

        if sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        if len(self._workbook.sheetnames) <= 1:
            raise ValueError("Cannot delete the last sheet")

        logger.info("Deleting sheet: %s", sheet_name)
        del self._workbook[sheet_name]

    def read_cell(self, sheet_name: str, cell: str) -> Any:
        """Read a single cell value."""
        ws = self.get_sheet(sheet_name)
        return ws[cell].value

    def write_cell(self, sheet_name: str, cell: str, value: Any) -> None:
        """Write a value to a cell."""
        ws = self.get_sheet(sheet_name)
        logger.debug("Writing %s to %s!%s", value, sheet_name, cell)
        ws[cell] = value

    def read_range(self, sheet_name: str, range_str: str) -> list[list[Any]]:
        """Read a range of cells."""
        ws = self.get_sheet(sheet_name)

        # Parse range (e.g., "A1:C10")
        start_cell, end_cell = range_str.split(":")

        # Convert to row/col numbers
        start_col = column_index_from_string("".join(c for c in start_cell if c.isalpha()))
        start_row = int("".join(c for c in start_cell if c.isdigit()))
        end_col = column_index_from_string("".join(c for c in end_cell if c.isalpha()))
        end_row = int("".join(c for c in end_cell if c.isdigit()))

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_ref = f"{get_column_letter(col)}{row}"
                row_data.append(ws[cell_ref].value)
            data.append(row_data)

        return data

    def write_range(self, sheet_name: str, range_str: str, values: list[list[Any]]) -> None:
        """Write values to a range."""
        ws = self.get_sheet(sheet_name)

        # Parse range
        start_cell = range_str.split(":")[0]

        # Convert to row/col numbers
        start_col = column_index_from_string("".join(c for c in start_cell if c.isalpha()))
        start_row = int("".join(c for c in start_cell if c.isdigit()))

        logger.debug("Writing %d rows to %s!%s", len(values), sheet_name, range_str)

        for i, row in enumerate(values):
            for j, value in enumerate(row):
                cell_ref = f"{get_column_letter(start_col + j)}{start_row + i}"
                ws[cell_ref] = value

    def get_used_range(self, sheet_name: str) -> str:
        """Get the used range of a worksheet."""
        ws = self.get_sheet(sheet_name)

        if ws.max_row is None or ws.max_column is None:
            return "A1"

        start = "A1"
        end = f"{get_column_letter(ws.max_column)}{ws.max_row}"

        return f"{start}:{end}"

    def get_cell_type(self, sheet_name: str, cell: str) -> str:
        """Get the data type of a cell."""
        ws = self.get_sheet(sheet_name)
        cell_obj = ws[cell]

        if cell_obj.value is None:
            return "null"

        # Check for date/datetime
        from datetime import date, datetime
        if isinstance(cell_obj.value, (datetime, date)):
            return "date"

        # Check for boolean
        if isinstance(cell_obj.value, bool):
            return "boolean"

        # Check for number
        if isinstance(cell_obj.value, (int, float)):
            return "number"

        # Default to string
        return "string"
