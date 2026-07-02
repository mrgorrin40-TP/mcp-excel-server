"""MCP tools for inspecting Excel files."""

import contextlib
import logging
from typing import Annotated, Any

from fastmcp import FastMCP
from pydantic import Field

from ..config import settings
from ..utils.backend import get_backend

logger = logging.getLogger(__name__)

# Create tools router
tools = FastMCP("Excel Inspection Tools", mask_error_details=True)


@tools.tool(
    name="get_column_stats",
    description="Get statistical summary of a column",
    tags={"excel", "read", "analysis"},
)
async def get_column_stats(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    column: Annotated[str, Field(description="Column letter or name")],
) -> dict[str, Any]:
    """Get statistical summary of a column."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        # Convert column letter to index if needed
        if column.isalpha():
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(column)
        else:
            col_idx = int(column)

        # Collect values
        values = []
        null_count = 0

        for row in range(2, (ws.max_row or 0) + 1):  # Skip header
            from openpyxl.utils import get_column_letter
            cell_ref = f"{get_column_letter(col_idx)}{row}"
            cell_value = ws[cell_ref].value

            if cell_value is None:
                null_count += 1
            elif isinstance(cell_value, int | float):
                values.append(cell_value)
            else:
                # Try to convert to number
                try:
                    values.append(float(cell_value))
                except (ValueError, TypeError):
                    null_count += 1

        # Calculate statistics
        if values:
            import statistics

            stats = {
                "count": len(values),
                "sum": sum(values),
                "mean": statistics.mean(values),
                "median": statistics.median(values),
                "min": min(values),
                "max": max(values),
                "range": max(values) - min(values),
                "std_dev": statistics.stdev(values) if len(values) > 1 else 0,
                "variance": statistics.variance(values) if len(values) > 1 else 0,
                "nulls": null_count,
                "type": "number",
            }
        else:
            stats = {
                "count": 0,
                "sum": 0,
                "mean": 0,
                "median": 0,
                "min": 0,
                "max": 0,
                "range": 0,
                "std_dev": 0,
                "variance": 0,
                "nulls": null_count,
                "type": "no_numeric_data",
            }

        return {
            "success": True,
            "column": column,
            "sheet_name": sheet_name,
            **stats,
        }
    except Exception as e:
        logger.error("Error getting column stats: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="filter_rows",
    description="Filter rows based on conditions",
    tags={"excel", "read", "analysis"},
)
async def filter_rows(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    filters: Annotated[list[dict[str, Any]], Field(description="Filter conditions")],
    logic: Annotated[str, Field(description="Logic operator: 'AND' or 'OR'")] = "AND",
) -> dict[str, Any]:
    """Filter rows based on conditions."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        # Get headers
        headers = []
        for col in range(1, (ws.max_column or 0) + 1):
            from openpyxl.utils import get_column_letter
            cell = ws[f"{get_column_letter(col)}1"]
            headers.append(str(cell.value) if cell.value else f"Column {col}")

        # Filter rows
        filtered_rows = []

        for row in range(2, (ws.max_row or 0) + 1):
            row_data = {}
            for col in range(1, (ws.max_column or 0) + 1):
                from openpyxl.utils import get_column_letter
                cell_ref = f"{get_column_letter(col)}{row}"
                row_data[headers[col - 1]] = ws[cell_ref].value

            # Check filters
            filter_results = []
            for f in filters:
                col_name = f.get("column")
                operator = f.get("operator", "==")
                value = f.get("value")

                if col_name not in row_data:
                    filter_results.append(False)
                    continue

                cell_value = row_data[col_name]

                # Apply operator
                if operator == "==":
                    result = cell_value == value
                elif operator == "!=":
                    result = cell_value != value
                elif operator == ">":
                    result = cell_value > value if cell_value is not None else False
                elif operator == "<":
                    result = cell_value < value if cell_value is not None else False
                elif operator == ">=":
                    result = cell_value >= value if cell_value is not None else False
                elif operator == "<=":
                    result = cell_value <= value if cell_value is not None else False
                elif operator == "contains":
                    result = (
                        str(value).lower() in str(cell_value).lower()
                        if cell_value is not None and value is not None
                        else False
                    )
                elif operator == "startswith":
                    result = str(cell_value).startswith(str(value)) if cell_value else False
                elif operator == "endswith":
                    result = str(cell_value).endswith(str(value)) if cell_value else False
                elif operator == "in":
                    result = cell_value in value if isinstance(value, list) else False
                elif operator == "not_in":
                    result = cell_value not in value if isinstance(value, list) else True
                elif operator == "is_null":
                    result = cell_value is None
                elif operator == "is_not_null":
                    result = cell_value is not None
                else:
                    result = False

                filter_results.append(result)

            # Apply logic
            matches = all(filter_results) if logic.upper() == "AND" else any(filter_results)

            if matches:
                filtered_rows.append(list(row_data.values()))

        return {
            "success": True,
            "data": filtered_rows[:settings.max_response_rows],
            "row_count": len(filtered_rows),
            "total_rows": (ws.max_row or 0) - 1,
            "truncated": len(filtered_rows) > settings.max_response_rows,
        }
    except Exception as e:
        logger.error("Error filtering rows: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="group_by",
    description="Group data and apply aggregation",
    tags={"excel", "read", "analysis"},
)
async def group_by(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    columns: Annotated[list[str], Field(description="Columns to group by")],
    agg_column: Annotated[str, Field(description="Column to aggregate")],
    agg_func: Annotated[
        str, Field(description="Aggregation function: sum, mean, median, count, min, max")
    ],
) -> dict[str, Any]:
    """Group data and apply aggregation."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        # Get headers
        headers = []
        for col in range(1, (ws.max_column or 0) + 1):
            from openpyxl.utils import get_column_letter
            cell = ws[f"{get_column_letter(col)}1"]
            headers.append(str(cell.value) if cell.value else f"Column {col}")

        # Collect data
        groups: dict[tuple[Any, ...], list[float]] = {}

        for row in range(2, (ws.max_row or 0) + 1):
            row_data = {}
            for col in range(1, (ws.max_column or 0) + 1):
                from openpyxl.utils import get_column_letter
                cell_ref = f"{get_column_letter(col)}{row}"
                row_data[headers[col - 1]] = ws[cell_ref].value

            # Create group key
            group_key = tuple(row_data.get(col) for col in columns)

            if group_key not in groups:
                groups[group_key] = []

            # Get aggregation value
            agg_value = row_data.get(agg_column)
            if agg_value is not None:
                with contextlib.suppress(ValueError, TypeError):
                    groups[group_key].append(float(agg_value))

        # Apply aggregation
        results = []
        for group_key, values in groups.items():
            if not values:
                continue

            import statistics

            if agg_func == "sum":
                agg_result = sum(values)
            elif agg_func == "mean":
                agg_result = statistics.mean(values)
            elif agg_func == "median":
                agg_result = statistics.median(values)
            elif agg_func == "count":
                agg_result = len(values)
            elif agg_func == "min":
                agg_result = min(values)
            elif agg_func == "max":
                agg_result = max(values)
            else:
                agg_result = sum(values)

            result = {col: val for col, val in zip(columns, group_key, strict=False)}
            result[f"{agg_column}_{agg_func}"] = agg_result
            results.append(result)

        return {
            "success": True,
            "groups": results,
            "group_count": len(results),
            "agg_function": agg_func,
            "agg_column": agg_column,
        }
    except Exception as e:
        logger.error("Error grouping data: %s", e)
        return {"success": False, "error": str(e)}
