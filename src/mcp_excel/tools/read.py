"""MCP tools for reading Excel data."""

import logging
from typing import Annotated, Any

from fastmcp import FastMCP
from pydantic import Field

from ..backends.factory import create_backend
from ..utils.cache import shared_cache

logger = logging.getLogger(__name__)

# Create tools router
tools = FastMCP("Excel Reading Tools", mask_error_details=True)


@tools.tool(
    name="read_cell",
    description="Read a single cell value from an Excel file",
    tags={"excel", "read"},
)
async def read_cell(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    cell: Annotated[str, Field(description="Cell reference (e.g., 'A1', 'B2')")],
) -> dict[str, Any]:
    """Read a single cell value from an Excel file."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        value = backend.read_cell(sheet_name, cell)
        cell_type = backend.get_cell_type(sheet_name, cell)

        return {
            "success": True,
            "cell": cell,
            "value": value,
            "type": cell_type,
            "sheet_name": sheet_name,
        }
    except Exception as e:
        logger.error("Error reading cell: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="read_range",
    description="Read a range of cells from an Excel file",
    tags={"excel", "read"},
)
async def read_range(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    cell_range: Annotated[str, Field(description="Range notation (e.g., 'A1:C10')")],
    page: Annotated[int, Field(description="Page number (1-indexed)", ge=1)] = 1,
    page_size: Annotated[int, Field(description="Rows per page", ge=1, le=1000)] = 100,
) -> dict[str, Any]:
    """Read a range of cells from an Excel file."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        # Get data
        data = backend.read_range(sheet_name, cell_range)

        # Apply pagination
        total_rows = len(data)
        start_idx = (page - 1) * page_size
        end_idx = min(start_idx + page_size, total_rows)

        paginated_data = data[start_idx:end_idx] if start_idx < total_rows else []

        return {
            "success": True,
            "data": paginated_data,
            "range": cell_range,
            "sheet_name": sheet_name,
            "total_rows": total_rows,
            "page": page,
            "page_size": page_size,
            "total_pages": (total_rows + page_size - 1) // page_size,
        }
    except Exception as e:
        logger.error("Error reading range: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="get_sheet_info",
    description="Get metadata about a worksheet",
    tags={"excel", "read", "metadata"},
)
async def get_sheet_info(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
) -> dict[str, Any]:
    """Get metadata about a worksheet."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        # Get sheet
        ws = backend.get_sheet(sheet_name)

        # Get used range
        used_range = backend.get_used_range(sheet_name)

        # Get column headers (first row)
        headers = []
        if ws.max_row and ws.max_row >= 1:
            for col in range(1, (ws.max_column or 0) + 1):
                from openpyxl.utils import get_column_letter
                cell = ws[f"{get_column_letter(col)}1"]
                headers.append({
                    "name": str(cell.value) if cell.value else f"Column {col}",
                    "letter": get_column_letter(col),
                    "type": backend.get_cell_type(sheet_name, f"{get_column_letter(col)}1"),
                })

        # Get sample data (rows 2-6)
        sample_data = []
        if ws.max_row and ws.max_row >= 2:
            for row in range(2, min(7, ws.max_row + 1)):
                row_data = []
                for col in range(1, (ws.max_column or 0) + 1):
                    from openpyxl.utils import get_column_letter
                    cell = ws[f"{get_column_letter(col)}{row}"]
                    row_data.append(cell.value)
                sample_data.append(row_data)

        return {
            "success": True,
            "name": sheet_name,
            "columns": headers,
            "row_count": ws.max_row or 0,
            "column_count": ws.max_column or 0,
            "used_range": used_range,
            "sample_data": sample_data,
        }
    except Exception as e:
        logger.error("Error getting sheet info: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="search_cells",
    description="Search for values in an Excel file",
    tags={"excel", "read", "search"},
)
async def search_cells(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    query: Annotated[str, Field(description="Search term")],
    sheet_name: Annotated[
        str | None, Field(description="Specific sheet (searches all if omitted)")
    ] = None,
    max_results: Annotated[int, Field(description="Maximum results to return", ge=1, le=500)] = 50,
) -> dict[str, Any]:
    """Search for values in an Excel file."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        matches = []
        sheets_to_search = [sheet_name] if sheet_name else backend.get_sheet_names()

        for sheet in sheets_to_search:
            try:
                ws = backend.get_sheet(sheet)

                for row in range(1, (ws.max_row or 0) + 1):
                    for col in range(1, (ws.max_column or 0) + 1):
                        from openpyxl.utils import get_column_letter
                        cell_ref = f"{get_column_letter(col)}{row}"
                        cell_value = ws[cell_ref].value

                        if cell_value and query.lower() in str(cell_value).lower():
                            matches.append({
                                "sheet": sheet,
                                "cell": cell_ref,
                                "row": row,
                                "column": get_column_letter(col),
                                "value": cell_value,
                            })

                            if len(matches) >= max_results:
                                return {
                                    "success": True,
                                    "matches": matches,
                                    "total_matches": len(matches),
                                    "max_results_reached": True,
                                }
            except Exception as e:
                logger.warning("Error searching sheet %s: %s", sheet, e)
                continue

        return {
            "success": True,
            "matches": matches,
            "total_matches": len(matches),
            "max_results_reached": False,
        }
    except Exception as e:
        logger.error("Error searching cells: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="list_sheets",
    description="List all worksheets in an Excel file",
    tags={"excel", "read", "metadata"},
)
async def list_sheets(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
) -> dict[str, Any]:
    """List all worksheets in an Excel file."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        sheet_names = backend.get_sheet_names()

        return {
            "success": True,
            "sheets": sheet_names,
            "count": len(sheet_names),
        }
    except Exception as e:
        logger.error("Error listing sheets: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="describe_workbook",
    description="Get comprehensive overview of an Excel workbook",
    tags={"excel", "read", "metadata"},
)
async def describe_workbook(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
) -> dict[str, Any]:
    """Get comprehensive overview of an Excel workbook."""
    try:
        backend = shared_cache.get(file_path)
        if backend is None:
            backend = create_backend(file_path)
            backend.open(file_path)
            shared_cache.put(file_path, backend)

        from pathlib import Path
        path = Path(file_path)

        sheet_names = backend.get_sheet_names()
        sheets_info = []

        for sheet_name in sheet_names:
            try:
                used_range = backend.get_used_range(sheet_name)
                ws = backend.get_sheet(sheet_name)

                sheets_info.append({
                    "name": sheet_name,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                    "used_range": used_range,
                })
            except Exception as e:
                logger.warning("Error getting info for sheet %s: %s", sheet_name, e)
                sheets_info.append({
                    "name": sheet_name,
                    "error": str(e),
                })

        return {
            "success": True,
            "file": {
                "name": path.name,
                "path": str(path.resolve()),
                "size_kb": round(path.stat().st_size / 1024, 2) if path.exists() else 0,
            },
            "sheets": sheets_info,
            "total_sheets": len(sheets_info),
        }
    except Exception as e:
        logger.error("Error describing workbook: %s", e)
        return {"success": False, "error": str(e)}
