"""MCP tools for Excel structured tables."""

import logging
from typing import Annotated, Any

from fastmcp import FastMCP
from pydantic import Field

from ..utils.backend import get_backend
from ..utils.common import col_letter_to_index

logger = logging.getLogger(__name__)

# Create tools router
tools = FastMCP("Excel Table Tools", mask_error_details=True)

# Table style options
TABLE_STYLES = [
    "TableStyleMedium2",
    "TableStyleMedium3",
    "TableStyleMedium4",
    "TableStyleMedium5",
    "TableStyleMedium6",
    "TableStyleMedium7",
    "TableStyleMedium8",
    "TableStyleMedium9",
    "TableStyleMedium10",
    "TableStyleLight1",
    "TableStyleLight2",
    "TableStyleLight3",
    "TableStyleLight4",
    "TableStyleLight5",
    "TableStyleLight6",
    "TableStyleLight7",
    "TableStyleLight8",
    "TableStyleLight9",
    "TableStyleLight10",
    "TableStyleDark1",
    "TableStyleDark2",
    "TableStyleDark3",
    "TableStyleDark4",
    "TableStyleDark5",
    "TableStyleDark6",
    "TableStyleDark7",
    "TableStyleDark8",
    "TableStyleDark9",
    "TableStyleDark10",
]


@tools.tool(
    name="create_table",
    description="Create a structured Excel table from a data range",
    tags={"excel", "write", "table"},
)
async def create_table(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    table_name: Annotated[str, Field(description="Name for the table")],
    data_range: Annotated[str, Field(description="Data range (e.g., 'A1:D100')")],
    has_header: Annotated[bool, Field(description="Whether data has headers")] = True,
    style: Annotated[
        str | None, Field(description="Table style (e.g., 'TableStyleMedium2')")
    ] = None,
) -> dict[str, Any]:
    """Create a structured Excel table from a data range."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        # Import table classes
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Parse data range
        start, end = data_range.split(":")
        col_letter_to_index(start)
        int("".join(c for c in start if c.isdigit()))
        col_letter_to_index(end)
        int("".join(c for c in end if c.isdigit()))

        # Create table
        tab = Table(displayName=table_name, ref=data_range)

        # Set style
        if style and style in TABLE_STYLES:
            tab.tableStyleInfo = TableStyleInfo(
                name=style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        else:
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

        # Add table to worksheet
        ws.add_table(tab)

        backend.save()

        return {
            "success": True,
            "table_name": table_name,
            "sheet_name": sheet_name,
            "data_range": data_range,
            "has_header": has_header,
            "style": style or "TableStyleMedium2",
        }
    except Exception as e:
        logger.error("Error creating table: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="list_tables",
    description="List all structured tables in an Excel worksheet",
    tags={"excel", "read", "table"},
)
async def list_tables(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
) -> dict[str, Any]:
    """List all structured tables in an Excel worksheet."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        tables = []
        # ws.tables is a TableList, iterating gives us table names
        for table_name in ws.tables:
            try:
                table = ws.tables[table_name]
                display_name = table.displayName if hasattr(table, "displayName") else table_name
                has_style = hasattr(table, "tableStyleInfo") and table.tableStyleInfo
                style_name = table.tableStyleInfo.name if has_style else None
                tables.append({
                    "name": table_name,
                    "display_name": display_name,
                    "ref": str(table.ref),
                    "style": style_name,
                })
            except Exception as e:
                logger.warning("Error reading table %s: %s", table_name, e)

        return {
            "success": True,
            "sheet_name": sheet_name,
            "tables": tables,
            "count": len(tables),
        }
    except Exception as e:
        logger.error("Error listing tables: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="delete_table",
    description="Delete a structured table from an Excel worksheet",
    tags={"excel", "write", "table"},
)
async def delete_table(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    table_name: Annotated[str, Field(description="Name of table to delete")],
) -> dict[str, Any]:
    """Delete a structured table from an Excel worksheet."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        if table_name not in ws.tables:
            return {
                "success": False,
                "error": f"Table '{table_name}' not found",
            }

        # Delete table
        del ws.tables[table_name]

        backend.save()

        return {
            "success": True,
            "sheet_name": sheet_name,
            "deleted_table": table_name,
        }
    except Exception as e:
        logger.error("Error deleting table: %s", e)
        return {"success": False, "error": str(e)}


@tools.tool(
    name="add_table_row",
    description="Add a row to an Excel structured table",
    tags={"excel", "write", "table"},
)
async def add_table_row(
    file_path: Annotated[str, Field(description="Absolute path to Excel file")],
    sheet_name: Annotated[str, Field(description="Worksheet name")],
    table_name: Annotated[str, Field(description="Table name")],
    values: Annotated[list[Any], Field(description="Row values")],
) -> dict[str, Any]:
    """Add a row to an Excel structured table."""
    try:
        backend = get_backend(file_path)

        ws = backend.get_sheet(sheet_name)

        if table_name not in ws.tables:
            return {
                "success": False,
                "error": f"Table '{table_name}' not found",
            }

        table = ws.tables[table_name]

        # Get current table range and expand by one row
        from openpyxl.utils import column_index_from_string, get_column_letter

        ref = str(table.ref)
        start, end = ref.split(":")

        # Parse start and end
        start_col = "".join(c for c in start if c.isalpha())
        start_row = int("".join(c for c in start if c.isdigit()))
        end_col = "".join(c for c in end if c.isalpha())
        end_row = int("".join(c for c in end if c.isdigit()))

        # Expand by one row
        new_end_row = end_row + 1
        new_ref = f"{start_col}{start_row}:{end_col}{new_end_row}"
        table.ref = new_ref

        # Write values to the new row
        end_col_index = column_index_from_string(end_col)
        for i, value in enumerate(values):
            col_index = column_index_from_string(start_col) + i
            if col_index <= end_col_index:
                cell = ws[f"{get_column_letter(col_index)}{new_end_row}"]
                cell.value = value

        backend.save()

        return {
            "success": True,
            "sheet_name": sheet_name,
            "table_name": table_name,
            "values": values,
        }
    except Exception as e:
        logger.error("Error adding table row: %s", e)
        return {"success": False, "error": str(e)}
