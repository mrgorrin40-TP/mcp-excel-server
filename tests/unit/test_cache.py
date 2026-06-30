"""Unit tests for LRU cache."""

import pytest
import time
from pathlib import Path

from mcp_excel.utils.cache import WorkbookCache
from mcp_excel.backends.openpyxl_backend import OpenpyxlBackend


class TestWorkbookCache:
    """Tests for WorkbookCache."""

    def test_put_and_get(self, sample_excel_file: Path):
        """Test basic put and get operations."""
        cache = WorkbookCache(max_size=10)
        backend = OpenpyxlBackend()
        backend.open(sample_excel_file)
        cache.put(sample_excel_file, backend)
        retrieved = cache.get(sample_excel_file)
        assert retrieved is not None
        cache.clear()

    def test_get_nonexistent_key(self):
        """Test getting a non-existent key returns None."""
        cache = WorkbookCache(max_size=10)
        assert cache.get("/nonexistent/file.xlsx") is None

    def test_eviction_when_full(self, sample_excel_file: Path, tmp_path: Path):
        """Test that oldest entry is evicted when cache is full."""
        cache = WorkbookCache(max_size=2)

        # Create second file
        from openpyxl import Workbook
        file2 = tmp_path / "test2.xlsx"
        wb = Workbook()
        wb.save(file2)

        backend1 = OpenpyxlBackend()
        backend1.open(sample_excel_file)
        cache.put(sample_excel_file, backend1)

        backend2 = OpenpyxlBackend()
        backend2.open(file2)
        cache.put(file2, backend2)

        # Create third file to trigger eviction
        file3 = tmp_path / "test3.xlsx"
        wb3 = Workbook()
        wb3.save(file3)
        backend3 = OpenpyxlBackend()
        backend3.open(file3)
        cache.put(file3, backend3)

        # First file should be evicted
        assert cache.get(sample_excel_file) is None
        cache.clear()

    def test_lru_order_update(self, sample_excel_file: Path, tmp_path: Path):
        """Test that accessing a key updates LRU order."""
        cache = WorkbookCache(max_size=2)

        from openpyxl import Workbook
        file2 = tmp_path / "test2.xlsx"
        wb = Workbook()
        wb.save(file2)

        file3 = tmp_path / "test3.xlsx"
        wb3 = Workbook()
        wb3.save(file3)

        backend1 = OpenpyxlBackend()
        backend1.open(sample_excel_file)
        cache.put(sample_excel_file, backend1)

        backend2 = OpenpyxlBackend()
        backend2.open(file2)
        cache.put(file2, backend2)

        # Access file1 to make it recently used
        cache.get(sample_excel_file)

        # Add file3 - should evict file2 (least recently used)
        backend3 = OpenpyxlBackend()
        backend3.open(file3)
        cache.put(file3, backend3)

        assert cache.get(sample_excel_file) is not None
        assert cache.get(file2) is None
        cache.clear()

    def test_clear(self, sample_excel_file: Path):
        """Test clearing the cache."""
        cache = WorkbookCache(max_size=10)
        backend = OpenpyxlBackend()
        backend.open(sample_excel_file)
        cache.put(sample_excel_file, backend)
        cache.clear()
        assert cache.size == 0

    def test_size(self, sample_excel_file: Path):
        """Test cache size tracking."""
        cache = WorkbookCache(max_size=10)
        assert cache.size == 0
        backend = OpenpyxlBackend()
        backend.open(sample_excel_file)
        cache.put(sample_excel_file, backend)
        assert cache.size == 1
        cache.clear()

    def test_remove(self, sample_excel_file: Path):
        """Test removing a specific key."""
        cache = WorkbookCache(max_size=10)
        backend = OpenpyxlBackend()
        backend.open(sample_excel_file)
        cache.put(sample_excel_file, backend)
        cache.remove(sample_excel_file)
        assert cache.get(sample_excel_file) is None
        assert cache.size == 0

    def test_keys(self, sample_excel_file: Path):
        """Test getting cache keys."""
        cache = WorkbookCache(max_size=10)
        backend = OpenpyxlBackend()
        backend.open(sample_excel_file)
        cache.put(sample_excel_file, backend)
        assert len(cache.keys) == 1
        cache.clear()
