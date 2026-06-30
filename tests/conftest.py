"""Pytest configuration and fixtures for MCP Excel Server tests."""

import pytest
from pathlib import Path
from typing import Generator


@pytest.fixture
def sample_excel_file(tmp_path: Path) -> Path:
    """Create a sample Excel file with data for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add headers
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"
    ws["D1"] = "Salary"

    # Add data rows
    ws["A2"] = "Alice"
    ws["B2"] = 25
    ws["C2"] = "New York"
    ws["D2"] = 50000

    ws["A3"] = "Bob"
    ws["B3"] = 30
    ws["C3"] = "London"
    ws["D3"] = 60000

    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Paris"
    ws["D4"] = 70000

    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def empty_excel_file(tmp_path: Path) -> Path:
    """Create an empty Excel file for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    file_path = tmp_path / "empty.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_sheet_excel_file(tmp_path: Path) -> Path:
    """Create an Excel file with multiple sheets."""
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Sales
    ws_sales = wb.active
    ws_sales.title = "Sales"
    ws_sales["A1"] = "Product"
    ws_sales["B1"] = "Quantity"
    ws_sales["C1"] = "Price"
    ws_sales["A2"] = "Widget"
    ws_sales["B2"] = 100
    ws_sales["C2"] = 10
    ws_sales["A3"] = "Gadget"
    ws_sales["B3"] = 50
    ws_sales["C3"] = 20

    # Sheet 2: Employees
    ws_emp = wb.create_sheet("Employees")
    ws_emp["A1"] = "Name"
    ws_emp["B1"] = "Department"
    ws_emp["A2"] = "Alice"
    ws_emp["B2"] = "Engineering"
    ws_emp["A3"] = "Bob"
    ws_emp["B2"] = "Marketing"

    file_path = tmp_path / "multi_sheet.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def large_excel_file(tmp_path: Path) -> Path:
    """Create a large Excel file for performance testing."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "LargeData"

    # Add headers
    ws["A1"] = "ID"
    ws["B1"] = "Value"
    ws["C1"] = "Category"

    # Add 1000 rows
    for i in range(2, 1002):
        ws[f"A{i}"] = i - 1
        ws[f"B{i}"] = i * 10
        ws[f"C{i}"] = f"Category_{i % 5}"

    file_path = tmp_path / "large.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def invalid_file(tmp_path: Path) -> Path:
    """Create an invalid file (not Excel) for error testing."""
    file_path = tmp_path / "invalid.txt"
    file_path.write_text("This is not an Excel file")
    return file_path
